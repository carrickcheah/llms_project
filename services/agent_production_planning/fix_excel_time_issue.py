#!/usr/bin/env python3
"""
Excel Time Extraction Fix using Epoch Timestamps

This script solves the problem of time information being lost from Excel date columns
by directly accessing the original Excel cells and converting dates to epoch timestamps.
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import logging
import os

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def analyze_excel_structure(file_path, sheet_name=None):
    """Analyze Excel structure to find the true header row and data fields."""
    # Load Excel file 
    if sheet_name is None:
        xl = pd.ExcelFile(file_path)
        sheet_name = xl.sheet_names[0]
        logger.info(f"Using first sheet: {sheet_name}")
    
    # Load all rows to analyze structure
    df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    logger.info(f"Excel file shape: {df_raw.shape}")
    
    # Look for key header indicators
    date_keywords = ["LCD_DATE", "START_DATE", "START DATE", "JOB", "PROCESS_CODE", "PROCESS CODE"]
    
    header_rows = []
    for row_idx in range(min(10, df_raw.shape[0])):  # Check first 10 rows
        row_values = df_raw.iloc[row_idx].astype(str)
        matches = sum(1 for val in row_values if any(kw in str(val).upper() for kw in date_keywords))
        if matches >= 2:  # Consider it a header row if multiple matches
            header_rows.append(row_idx)
            logger.info(f"Possible header at row {row_idx+1}: {matches} keyword matches")
            logger.info(f"Sample values: {row_values.head(5).tolist()}")
    
    # Also check for the rows that have date values
    date_cells = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    for row_idx, row in enumerate(sheet.rows):
        for col_idx, cell in enumerate(row):
            if isinstance(cell.value, datetime):
                date_cells.append((row_idx, col_idx, cell.value))
    
    logger.info(f"Found {len(date_cells)} date cells in Excel")
    if date_cells:
        # Sample a few dates
        for i, (row, col, value) in enumerate(date_cells[:5]):
            logger.info(f"Date at row {row+1}, col {sheet.cell(row=1, column=col+1).column_letter}: {value}")
    
    # Try to figure out the best header row
    best_header_row = None
    if header_rows:
        # Prefer the first row that has keyword matches
        best_header_row = header_rows[0]
    
    return {
        'sheet_name': sheet_name,
        'best_header_row': best_header_row,
        'header_rows': header_rows,
        'date_cells': date_cells
    }

def process_excel_with_time(file_path, sheet_name=None, header_row=None):
    """
    Process Excel file preserving time components of dates by using epoch timestamps.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to process (None for first sheet)
        header_row: Row number containing headers (0-based, None for auto-detect)
        
    Returns:
        DataFrame with epoch timestamp columns
    """
    # First analyze Excel structure if needed
    if header_row is None or sheet_name is None:
        analysis = analyze_excel_structure(file_path, sheet_name)
        sheet_name = analysis['sheet_name']
        header_row = analysis['best_header_row'] if header_row is None else header_row
    
    logger.info(f"Processing Excel with sheet: {sheet_name}, header row: {header_row+1 if header_row is not None else 'auto'}")
    
    # Now read the Excel file
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    logger.info(f"Loaded DataFrame with {df.shape[0]} rows and {df.shape[1]} columns")
    
    # Display column information
    logger.info("Column names in DataFrame:")
    for i, col in enumerate(df.columns):
        logger.info(f"{i+1}. '{col}' (type: {type(col).__name__})")
    
    # Look for date columns
    date_columns = []
    for col in df.columns:
        col_str = str(col).upper()
        if "DATE" in col_str or "START" in col_str or "END" in col_str:
            date_columns.append(col)
    
    logger.info(f"Found {len(date_columns)} potential date columns: {date_columns}")
    
    # Extract datetime values directly from Excel cells
    logger.info("Loading workbook to extract raw date and time values...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    # Find column indices in the Excel worksheet
    excel_header_row = header_row + 1  # Convert to 1-based for openpyxl
    column_indices = {}
    
    # Map DataFrame columns to Excel column indices
    header_cells = list(ws.rows)[excel_header_row]
    for i, cell in enumerate(header_cells, 1):  # 1-based column indexing in openpyxl
        cell_value = cell.value
        if cell_value:
            for col in df.columns:
                if str(cell_value).strip() == str(col).strip():
                    column_indices[col] = i
                    break
    
    logger.info(f"Column index mapping: {column_indices}")
    
    # Extract date cells and their time components
    date_values = {}
    first_data_row = excel_header_row + 1  # 1-based row indexing
    
    # Process all date columns
    for col in date_columns:
        if col in column_indices:
            col_idx = column_indices[col]
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            logger.info(f"Extracting dates from column {col} (Excel column {col_letter})")
            
            # Create epoch timestamp column
            epoch_col = f"epoch_{col.lower().replace(' ', '_').replace('-', '_')}"
            df[epoch_col] = np.nan
            
            # Counters for statistics
            total_dates = 0
            with_time = 0
            
            # Find the corresponding index in the pandas DataFrame
            for excel_row in range(first_data_row, first_data_row + df.shape[0]):
                try:
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if isinstance(cell.value, datetime):
                        # Convert to pandas row index (0-based)
                        pandas_row = excel_row - first_data_row
                        
                        if pandas_row < df.shape[0]:  # Ensure we're within DataFrame bounds
                            # Get the full datetime value with time component
                            dt = cell.value
                            
                            # Convert to epoch timestamp (seconds since Unix epoch)
                            epoch_timestamp = int(dt.timestamp())
                            
                            # Store in DataFrame
                            df.loc[pandas_row, epoch_col] = epoch_timestamp
                            
                            # Update statistics
                            total_dates += 1
                            if dt.hour != 0 or dt.minute != 0 or dt.second != 0:
                                with_time += 1
                                
                            # Debug log the first few values
                            if total_dates <= 5:
                                logger.info(f"Row {excel_row}: {dt.strftime('%Y-%m-%d %H:%M:%S')} → Epoch: {epoch_timestamp}")
                except Exception as e:
                    logger.error(f"Error processing cell at row {excel_row}, column {col_idx}: {e}")
            
            logger.info(f"Processed {total_dates} dates in column '{col}', {with_time} had non-zero time components")
            
            # For cells that don't have dates in Excel, try to parse from string values
            missing_mask = df[epoch_col].isna()
            if missing_mask.any():
                logger.info(f"{missing_mask.sum()} rows without Excel date objects in '{col}' - trying string parsing")
                
                # Try multiple date formats
                for fmt in ['%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M', '%m/%d/%Y %H:%M', 
                            '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        # Parse dates, keeping NaN for rows that already have timestamps
                        parsed = pd.to_datetime(df.loc[missing_mask, col], format=fmt, errors='coerce')
                        valid = parsed.notna().sum()
                        
                        if valid > 0:
                            logger.info(f"Parsed {valid} additional dates using format '{fmt}'")
                            
                            # Convert to epoch timestamps and update the DataFrame
                            for idx, dt in parsed.items():
                                if pd.notna(dt):
                                    # Convert datetime to epoch timestamp
                                    df.loc[idx, epoch_col] = int(dt.timestamp())
                            
                            break  # Stop after finding a working format
                    except Exception as e:
                        continue
        else:
            logger.warning(f"Column '{col}' not found in Excel column mapping")
    
    # Also add a shortcut column for START_DATE specifically, if it exists
    if any('START_DATE' in str(col).upper() for col in df.columns):
        # Find the actual column name that contains START_DATE
        start_date_col = next((col for col in df.columns if 'START_DATE' in str(col).upper()), None)
        if start_date_col and f"epoch_{start_date_col.lower().replace(' ', '_').replace('-', '_')}" in df.columns:
            epoch_col = f"epoch_{start_date_col.lower().replace(' ', '_').replace('-', '_')}"
            df['start_timestamp'] = df[epoch_col]
            logger.info(f"Added 'start_timestamp' column as a copy of {epoch_col}")
    
    return df

def run_time_fix(input_file=None, output_file=None):
    """Run the Excel time fix and save the results."""
    # Use default paths if not provided
    if input_file is None:
        input_file = "/Users/carrickcheah/llms_project/services/agent_production_planning/mydata.xlsx"
    
    if output_file is None:
        # Create output filename based on input
        base_dir = os.path.dirname(input_file)
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(base_dir, f"{base_name}_with_epochs.xlsx")
    
    logger.info(f"Processing file: {input_file}")
    logger.info(f"Output will be saved to: {output_file}")
    
    # Process the Excel file
    result_df = process_excel_with_time(input_file)
    
    # Count epoch columns
    epoch_cols = [col for col in result_df.columns if col.startswith('epoch_')]
    logger.info(f"Created {len(epoch_cols)} epoch timestamp columns: {epoch_cols}")
    
    # Show sample data with timestamps converted back to datetime for verification
    if epoch_cols:
        logger.info("\nSample data with epoch timestamps converted back to datetime:")
        sample_rows = min(5, result_df.shape[0])
        for epoch_col in epoch_cols:
            sample = result_df[epoch_col].dropna().head(sample_rows)
            if not sample.empty:
                logger.info(f"\nColumn: {epoch_col}")
                for idx, epoch_val in sample.items():
                    # Convert epoch back to datetime for display
                    dt = datetime.fromtimestamp(epoch_val)
                    logger.info(f"  Row {idx+1}: {epoch_val} → {dt.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Save the result
    result_df.to_excel(output_file, index=False)
    logger.info(f"Results saved to {output_file}")
    
    return result_df

if __name__ == "__main__":
    run_time_fix()
