#!/usr/bin/env python3
"""
Fix Time Disparency in Excel Date Columns

This script directly maps Excel date columns to epoch timestamps to ensure
time components are preserved. It specifically fixes the issue where time
information is lost when reading dates from Excel files.
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import logging
import os

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def fix_time_disparency(file_path, sheet_name=None):
    """
    Fix time disparency in Excel date columns by using epoch timestamps.
    Takes a direct approach to ensure time components are preserved.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to process (None for first sheet)
        
    Returns:
        DataFrame with added epoch timestamp columns
    """
    # Determine sheet name if not provided
    if sheet_name is None:
        xl = pd.ExcelFile(file_path)
        sheet_name = xl.sheet_names[0]
        logger.info(f"Using first sheet: {sheet_name}")
    
    logger.info(f"Processing Excel file: {file_path}, Sheet: {sheet_name}")
    
    # Load Excel with openpyxl to directly access cells and preserve time components
    logger.info("Loading Excel with openpyxl to preserve date/time information...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    # First, find the header row with key column names
    header_row = None
    date_keywords = ["LCD_DATE", "START_DATE", "JOB", "PROCESS_CODE"]
    
    # Search for header row in first 10 rows
    for row_idx in range(1, 11):
        row_values = [str(cell.value).upper() if cell.value else "" for cell in ws[row_idx]]
        matches = sum(1 for val in row_values if any(kw in val for kw in date_keywords))
        if matches >= 2:  # Found likely header row with multiple matches
            header_row = row_idx
            logger.info(f"Found header row at row {header_row}")
            logger.info(f"Header values: {[cell.value for cell in ws[row_idx] if cell.value]}")
            break
    
    if header_row is None:
        logger.error("Could not find header row. Using row 4 as default.")
        header_row = 4
    
    # Map header columns to indices
    header_map = {}
    date_columns = []
    
    for col_idx, cell in enumerate(ws[header_row], 1):  # 1-based column indexing
        if cell.value:
            col_name = str(cell.value).strip()
            header_map[col_name] = col_idx
            
            # Identify likely date columns
            if "DATE" in col_name.upper() or "START" in col_name.upper() or "END" in col_name.upper():
                date_columns.append(col_name)
    
    logger.info(f"Found {len(header_map)} header columns: {list(header_map.keys())}")
    logger.info(f"Identified {len(date_columns)} date columns: {date_columns}")
    
    # Now read Excel with pandas using the identified header row
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row-1)
    logger.info(f"Loaded DataFrame with {df.shape[0]} rows and {df.shape[1]} columns")
    
    # Process date columns and add epoch timestamp columns
    first_data_row = header_row + 1  # First row of data
    
    for col_name in date_columns:
        # Try to find the column in DataFrame (may have trailing spaces)
        df_col = None
        for df_col_name in df.columns:
            if str(df_col_name).strip() == col_name:
                df_col = df_col_name
                break
        
        if df_col is None:
            logger.warning(f"Column '{col_name}' not found in DataFrame columns")
            continue
        
        # Find the Excel column index
        if col_name not in header_map:
            logger.warning(f"Column '{col_name}' not found in Excel header mapping")
            continue
        
        excel_col_idx = header_map[col_name]
        logger.info(f"Processing date column: '{col_name}' (Excel column {ws.cell(row=1, column=excel_col_idx).column_letter})")
        
        # Create epoch timestamp column
        epoch_col = f"epoch_{col_name.lower().replace(' ', '_').replace('-', '_')}"
        df[epoch_col] = np.nan
        
        # Extract date cells with time components
        timestamps_found = 0
        with_time = 0
        
        for row_num in range(first_data_row, first_data_row + df.shape[0]):
            try:
                # Get cell value from Excel
                cell = ws.cell(row=row_num, column=excel_col_idx)
                
                # Convert to pandas DataFrame index (0-based)
                pandas_idx = row_num - first_data_row
                
                # Skip if we've reached the end of DataFrame
                if pandas_idx >= df.shape[0]:
                    break
                
                # Process if cell contains datetime
                if isinstance(cell.value, datetime):
                    # Get the datetime with its time component
                    dt = cell.value
                    
                    # Convert to epoch timestamp (seconds since Unix epoch)
                    epoch_timestamp = int(dt.timestamp())
                    
                    # Store in DataFrame
                    df.loc[pandas_idx, epoch_col] = epoch_timestamp
                    
                    # Update statistics
                    timestamps_found += 1
                    
                    # Count entries with non-zero time
                    if dt.hour != 0 or dt.minute != 0 or dt.second != 0:
                        with_time += 1
                    
                    # Show first few entries
                    if timestamps_found <= 5:
                        process = df.loc[pandas_idx, 'PROCESS_CODE'] if 'PROCESS_CODE' in df.columns else f"Row {pandas_idx+1}"
                        logger.info(f"Row {row_num}: {process} - {dt.strftime('%Y-%m-%d %H:%M:%S')} → Epoch: {epoch_timestamp}")
            except Exception as e:
                logger.error(f"Error processing cell at row {row_num}, column {excel_col_idx}: {e}")
        
        logger.info(f"Added {timestamps_found} timestamps to '{epoch_col}', {with_time} had non-zero time components")
        
        # For START_DATE, it's optional - we don't set default values
        if col_name == 'START_DATE' or col_name == 'START_DATE ':
            missing_mask = df[epoch_col].isna()
            missing_count = missing_mask.sum()
            logger.info(f"START_DATE is optional: {timestamps_found} values found, {missing_count} are empty and will remain empty")
    
    # Add helper column for START_DATE specifically
    start_date_epoch_cols = [col for col in df.columns if 'start_date' in col.lower() and 'epoch' in col.lower()]
    if start_date_epoch_cols:
        df['start_timestamp'] = df[start_date_epoch_cols[0]]
        logger.info(f"Added 'start_timestamp' column from {start_date_epoch_cols[0]}")
    
    # Count epoch columns
    epoch_cols = [col for col in df.columns if col.startswith('epoch_')]
    logger.info(f"Added {len(epoch_cols)} epoch timestamp columns: {epoch_cols}")
    
    # Show sample data with timestamps converted back to datetime
    if epoch_cols:
        logger.info("\nSample data with epoch timestamps converted back to datetime:")
        sample_rows = min(5, df.shape[0])
        for epoch_col in epoch_cols:
            sample = df[epoch_col].dropna().head(sample_rows)
            if not sample.empty:
                logger.info(f"\nColumn: {epoch_col}")
                for idx, epoch_val in sample.items():
                    # Convert epoch back to datetime for display
                    dt = datetime.fromtimestamp(epoch_val)
                    logger.info(f"  Row {idx+1}: {epoch_val} → {dt.strftime('%Y-%m-%d %H:%M:%S')}")
    
    return df

def run_fix(input_file=None, output_file=None):
    """Run the time disparency fix and save results."""
    # Use default paths if not provided
    if input_file is None:
        input_file = "/Users/carrickcheah/llms_project/services/agent_production_planning/mydata.xlsx"
    
    if output_file is None:
        # Create output filename
        base_dir = os.path.dirname(input_file)
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(base_dir, f"{base_name}_time_fixed.xlsx")
    
    logger.info(f"Input file: {input_file}")
    logger.info(f"Output will be saved to: {output_file}")
    
    # Process the file
    result_df = fix_time_disparency(input_file)
    
    # Save the result
    result_df.to_excel(output_file, index=False)
    logger.info(f"Results saved to {output_file}")
    
    return result_df

if __name__ == "__main__":
    run_fix()
