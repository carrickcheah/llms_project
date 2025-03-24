"""
Excel to Epoch Timestamp Converter

This module provides reliable extraction of date and time information from Excel files,
converting them to epoch timestamps (seconds since Unix epoch) to prevent time data loss.
"""

import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
import logging
import re
from ingest_config import DATE_COLUMNS_CONFIG, DEFAULT_TIME

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_excel_raw_data(file_path, sheet_name=0, header_row=3):
    """
    Load Excel data preserving raw cell values and formats using openpyxl.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name or index
        header_row: Row index containing column headers (0-based in pandas, 1-based in openpyxl)
        
    Returns:
        Dictionary containing:
            - 'pandas_df': DataFrame from pandas
            - 'column_formats': Dictionary mapping column names to Excel number formats
            - 'date_cells': Dictionary mapping (row, col) to datetime values directly from Excel
    """
    logger.info(f"Loading Excel file with pandas and openpyxl: {file_path}")
    
    # Load with pandas first to get column structure
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    
    # Now use openpyxl to get precise cell formats and values
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet = wb[sheet_name]
    
    # Get header row in openpyxl (add 1 to convert from pandas 0-index to openpyxl 1-index)
    openpyxl_header_row = header_row + 1
    header_cells = list(sheet.rows)[openpyxl_header_row]
    
    # Map column indices to names
    col_idx_to_name = {}
    for idx, cell in enumerate(header_cells):
        if cell.value:
            col_name = str(cell.value).strip()
            col_idx_to_name[idx+1] = col_name  # 1-based indexing in openpyxl
    
    # Get column formats
    column_formats = {}
    for col_idx, col_name in col_idx_to_name.items():
        # Check first few non-empty cells to determine format
        for row_idx in range(openpyxl_header_row + 1, openpyxl_header_row + 10):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                column_formats[col_name] = cell.number_format
                break
    
    # Extract all date cells directly from openpyxl
    date_cells = {}
    date_column_indices = []
    
    # Identify potential date columns by name or format
    date_keywords = ['DATE', 'START', 'END', 'DUE', 'PLANNED', 'COMPLETION']
    
    for col_idx, col_name in col_idx_to_name.items():
        if any(keyword in col_name.upper() for keyword in date_keywords):
            date_column_indices.append(col_idx)
            logger.info(f"Identified potential date column: {col_name} (col {col_idx})")
    
    # Also identify columns with date formats
    for col_idx, col_name in col_idx_to_name.items():
        if col_idx not in date_column_indices and col_name in column_formats:
            fmt = column_formats[col_name]
            # Check for common date format codes in Excel
            if any(code in fmt for code in ['yy', 'mm', 'dd', 'm/d', 'd/m', 'mmm']):
                date_column_indices.append(col_idx)
                logger.info(f"Identified date column by format: {col_name} (format: {fmt})")
    
    # Extract date values from these columns
    for col_idx in date_column_indices:
        col_name = col_idx_to_name.get(col_idx, f"Column {col_idx}")
        
        for row_idx in range(openpyxl_header_row + 1, openpyxl_header_row + len(df) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # If the cell contains a datetime value
            if isinstance(cell.value, datetime):
                # Store with pandas row index (0-based)
                pandas_row_idx = row_idx - openpyxl_header_row - 1
                if pandas_row_idx < len(df):  # Ensure we're within pandas DataFrame bounds
                    date_cells[(pandas_row_idx, col_name)] = cell.value
                    
                    # Debug first few cells
                    if pandas_row_idx < 5:
                        logger.info(f"Found date in {col_name}, row {pandas_row_idx+1}: "
                                   f"{cell.value.strftime('%Y-%m-%d %H:%M:%S')}")
    
    return {
        'pandas_df': df,
        'column_formats': column_formats,
        'date_cells': date_cells
    }

def extract_time_from_string(value):
    """Extract time (hours, minutes) from string if present."""
    if not isinstance(value, str):
        return None
    
    # Try various time patterns
    patterns = [
        r'(\d{1,2}):(\d{2})(?::(\d{2}))?',  # HH:MM or HH:MM:SS
        r'(\d{1,2})\.(\d{2})',               # HH.MM
        r'(\d{1,2})h(\d{2})',                # HHhMM
    ]
    
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            groups = match.groups()
            hour = int(groups[0])
            minute = int(groups[1])
            # Ensure valid hours and minutes
            if 0 <= hour <= 23 and 0 <= minute <= 59:
                return (hour, minute)
    
    return None

def normalize_date_columns(data, date_columns=None):
    """
    Convert all date columns to epoch timestamps, preserving time components.
    
    Args:
        data: Dictionary from get_excel_raw_data
        date_columns: List of date column names to process, or None to auto-detect
        
    Returns:
        DataFrame with epoch timestamp columns added
    """
    df = data['pandas_df'].copy()
    date_cells = data['date_cells']
    
    # Auto-detect date columns if not specified
    if date_columns is None:
        # Use the date cells we extracted to determine date columns
        detected_date_columns = set()
        for _, col_name in date_cells.keys():
            detected_date_columns.add(col_name)
        
        # Also check for columns with 'date' in the name
        for col in df.columns:
            if 'DATE' in str(col).upper():
                detected_date_columns.add(col)
        
        date_columns = list(detected_date_columns)
    
    logger.info(f"Processing date columns: {date_columns}")
    
    for col in date_columns:
        if col not in df.columns:
            logger.warning(f"Column '{col}' not found in DataFrame")
            continue
        
        epoch_col = f"epoch_{col.lower().replace(' ', '_').replace('-', '_')}"
        logger.info(f"Converting column '{col}' to epoch timestamps in '{epoch_col}'")
        
        # Initialize epoch column
        df[epoch_col] = None
        
        # 1. First use the openpyxl-extracted datetime values (most reliable)
        cells_found = 0
        for (row_idx, cell_col), dt_value in date_cells.items():
            if cell_col == col and row_idx < len(df):
                # Use the extracted datetime with its original time component
                df.loc[row_idx, epoch_col] = int(dt_value.timestamp())
                cells_found += 1
        
        logger.info(f"Applied {cells_found} datetime values from openpyxl for '{col}'")
        
        # 2. For cells without openpyxl data, try pandas datetime conversion
        missing_mask = df[epoch_col].isna()
        if missing_mask.any():
            missing_count = missing_mask.sum()
            logger.info(f"Converting {missing_count} remaining values in '{col}' using pandas")
            
            try:
                # Convert to datetime using pandas
                dates = pd.to_datetime(df.loc[missing_mask, col], errors='coerce')
                
                # Apply the datetime values to the epoch column
                for idx, date_val in dates.items():
                    if pd.notna(date_val):
                        # Preserve the full datetime including time
                        df.loc[idx, epoch_col] = int(date_val.timestamp())
            except Exception as e:
                logger.error(f"Error converting values in '{col}' using pandas: {e}")
        
        # 3. For cells that are still missing, try to extract time from strings
        missing_mask = df[epoch_col].isna()
        if missing_mask.any():
            missing_count = missing_mask.sum()
            logger.info(f"Attempting to extract time from {missing_count} string values in '{col}'")
            
            for idx, value in df.loc[missing_mask, col].items():
                if isinstance(value, str):
                    # Try to extract a date using multiple formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                        try:
                            # First try to get just the date
                            date_obj = datetime.strptime(value.split()[0], fmt)
                            
                            # Then check if there's a time component in the string
                            time_info = extract_time_from_string(value)
                            
                            if time_info:
                                hour, minute = time_info
                                date_obj = date_obj.replace(hour=hour, minute=minute)
                            else:
                                # Use configured default time for this column
                                col_base = col.strip()
                                hour, minute = DATE_COLUMNS_CONFIG.get(col_base, DEFAULT_TIME)
                                date_obj = date_obj.replace(hour=hour, minute=minute)
                            
                            # Convert to epoch timestamp
                            df.loc[idx, epoch_col] = int(date_obj.timestamp())
                            break  # Stop after first successful parse
                        except:
                            continue
        
        # 4. Apply configured default times to any timestamps with midnight time
        col_base = col.strip()
        if col_base in DATE_COLUMNS_CONFIG:
            default_hour, default_minute = DATE_COLUMNS_CONFIG[col_base]
            
            # Identify cells with epoch timestamps that have midnight time (00:00)
            midnight_times = 0
            noon_count = 0
            
            for idx, epoch_val in df[epoch_col].items():
                if pd.notna(epoch_val):
                    # Convert epoch to datetime to check time component
                    dt = datetime.fromtimestamp(epoch_val)
                    
                    # Check if time is midnight
                    if dt.hour == 0 and dt.minute == 0:
                        # Replace with configured time
                        new_dt = dt.replace(hour=default_hour, minute=default_minute)
                        df.loc[idx, epoch_col] = int(new_dt.timestamp())
                        midnight_times += 1
                    # Check if the time is already our configured time
                    elif dt.hour == default_hour and dt.minute == default_minute:
                        noon_count += 1
            
            if midnight_times > 0:
                logger.info(f"Applied configured time ({default_hour}:{default_minute:02d}) to {midnight_times} "
                           f"midnight values in '{col}'")
            
            logger.info(f"{noon_count} values already had configured time {default_hour}:{default_minute:02d}")
        
        # Report valid timestamp count
        valid_count = df[epoch_col].notna().sum()
        logger.info(f"Final result: {valid_count}/{len(df)} valid timestamps for '{col}'")
        
        # Sample a few values for verification
        if valid_count > 0:
            sample = df[df[epoch_col].notna()].iloc[:5]
            for idx, row in sample.iterrows():
                epoch_val = row[epoch_col]
                dt = datetime.fromtimestamp(epoch_val)
                orig_val = row[col]
                logger.info(f"  Sample {col}: '{orig_val}' → {dt.strftime('%Y-%m-%d %H:%M:%S')} ({epoch_val})")
    
    return df

def excel_to_epoch_main(file_path, sheet_name=0, header_row=3, date_columns=None):
    """
    Main function to convert Excel date columns to epoch timestamps.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name or index
        header_row: Row index containing column headers (0-based)
        date_columns: List of date column names to process, or None to auto-detect
        
    Returns:
        DataFrame with added epoch timestamp columns for all date columns
    """
    # Extract raw Excel data
    data = get_excel_raw_data(file_path, sheet_name, header_row)
    
    # Normalize date columns to epoch timestamps
    df_with_epochs = normalize_date_columns(data, date_columns)
    
    return df_with_epochs

if __name__ == "__main__":
    # Example usage
    file_path = "/Users/carrickcheah/llms_project/services/agent_production_planning/mydata.xlsx"
    
    # Specify date columns to ensure we process all required ones
    date_cols = ['LCD_DATE', 'START_DATE', 'PLANNED_START', 'PLANNED_END', 'LATEST_COMPLETION_DATE']
    
    # Process the file
    result_df = excel_to_epoch_main(file_path, sheet_name='Jobs PlanningDetails-Draft', date_columns=date_cols)
    
    # Display results for each date column
    print("\nSummary of Epoch Conversions:")
    for col in date_cols:
        epoch_col = f"epoch_{col.lower().replace(' ', '_').replace('-', '_')}"
        if epoch_col in result_df.columns:
            valid_count = result_df[epoch_col].notna().sum()
            print(f"{col}: {valid_count}/{len(result_df)} values converted to epochs")
            
            # Display a few examples with their timestamp values
            if valid_count > 0:
                print("Sample values:")
                sample = result_df[result_df[epoch_col].notna()].head(3)
                for idx, row in sample.iterrows():
                    epoch_val = row[epoch_col]
                    dt = datetime.fromtimestamp(epoch_val)
                    orig_val = row[col] if col in row else "N/A"
                    print(f"  Row {idx+1}: '{orig_val}' → {dt.strftime('%Y-%m-%d %H:%M:%S')} (epoch: {epoch_val})")
    
    print("\nDone! Processed all date columns to epoch timestamps.")
